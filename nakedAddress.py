from openpyxl.styles import colors, PatternFill, Font, Color
from openpyxl import load_workbook
from openpyxl import Workbook
from datetime import datetime
from decimal import Decimal
import requests
import shutil
import json
import time
import os
import re

INPUT_PATH = 'input/'
OUTPUT_PATH = 'output/'

asset_usd_mapping = {
    'bitcoin': 'btcusd',
    'bitcoin cash': 'bchusd',
    'dogecoin': 'dogeusd',
    'dash': 'dashusd',
    'litecoin': 'ltcusd',
    'zcash': 'zecusd',
    'ethereum': 'ethusd',
    'binance smart chain': 'bnbusd',
    'polygon': 'maticusd',
    'polkadot': 'dotusd',
    'solana': 'solusd',
    'algorand': 'algousd',
    'ripple': 'xrpusd'
}

HEADERS = {
    "user-agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/74.0.3729.169 Safari/537.36" ,
    'referer':'https://www.google.com/'
}

BLOCKDAEMON_BEARER = 'zudFn9dhHe1lTgOeQ9B22MR6DHo-sUOIIRbBSg9GGyWwYAGQ'

ASSET_TX_URLS = {
    'BTC' : 'https://www.blockchain.com/explorer/transactions/btc/',
    'BCH' : 'https://www.blockchain.com/explorer/transactions/bch/',
    'DASH' : 'https://explorer.dash.org/insight/tx/',
    'DOGE' : 'https://dogechain.info/tx/',
    'LTC' : 'https://litecoinblockexplorer.net/tx/',
    'ZEC' : 'https://explorer.zcha.in/transactions/',
    'ALGO' : 'https://algoexplorer.io/tx/',
    'XRP' : 'https://xrpscan.com/tx/',
    'XLM' : 'https://stellarchain.io/transactions/',
    'ETC' : 'https://etcblockexplorer.com/tx/',
    'DOT' : 'https://polkadot.subscan.io/extrinsic/',
    'SOL' : 'https://explorer.solana.com/tx/'
}

def set_verified_tx(work_sheet, row, column, value):
    work_sheet[row][column].value = f'=HYPERLINK("{value}", "Verified")'
    work_sheet[row][column].fill = PatternFill(start_color='C3ECCB', end_color='C3ECCB', fill_type = 'solid')
    work_sheet[row][column].font = Font(color='006100')

def set_worksheet(output, asset):
    ws_count = len(output.sheetnames)
    work_sheet = output.active
    if ws_count == 1 and work_sheet.title == 'Sheet':
        work_sheet.title = asset
    else:
        output.create_sheet(asset)
        output.active = ws_count
        work_sheet = output.active
        
    return work_sheet
    
def get_blockexplorer_pagecount(api_link, address, tx_keys):
    link = f'{api_link}address/{address}'
    is_valid = requests.get(link, headers=HEADERS)
    
    if is_valid.status_code != 200:
        return 0
        
    content = json.loads(is_valid.content)
    
    tx_count = 0
    
    for tx_key in tx_keys:
        tx_count += content[tx_key]
        
    if tx_count == 0:
        return tx_count
    
    return content['totalPages']
        
def process_stellar(asset, asset_symbol, asset_decimal, address, start_date, end_date, output):
    api_link = f'https://horizon.stellar.org/accounts/{address}/transactions?limit=200'

    txns = requests.get(api_link, headers=HEADERS)
    content = json.loads(txns.content)

    work_sheet = set_worksheet(output, asset)
    work_sheet.append(['Date', 'Block Index', 'Transaction ID', 'Sent', 'Received', 'Asset', 'Coming From', 'Going To', 'Blockchain URL'])        
    
    rows = content['_embedded']['records']
    cursor = re.search(r'cursor=(\d+)', content['_links']['next']['href']).group(1)
    
    while True:
        for row in rows:
            timestamp = int(datetime.strptime(row['created_at'].split('T')[0], "%Y-%m-%d").timestamp())
            if timestamp < start_date:
                continue
            if timestamp > end_date:
                break
                
            tx_link = f'https://horizon.stellar.org/transactions/{row["id"]}/operations'
            tx_details = requests.get(tx_link, headers=HEADERS)
            tx_content = json.loads(tx_details.content)
            data = tx_content['_embedded']['records'][0]
            if data['type'] != 'payment':
                continue 
            sent = Decimal(data['amount']) if data['from'] == address else 0
            received = Decimal(data['amount']) if data['to'] == address else 0
            work_sheet.append([row['created_at'], row['ledger'], row['id'], sent, received, asset_symbol, data['from'], data['to'], ''])
            set_verified_tx(work_sheet, work_sheet.max_row, 8, f'{ASSET_TX_URLS[asset_symbol]}{row["id"]}')        
        
        txns = requests.get(f'{api_link}&cursor={cursor}', headers=HEADERS)
        content = json.loads(txns.content)
        rows = content['_embedded']['records']
        cursor = re.search(r'cursor=(\d+)', content['_links']['next']['href']).group(1)
        if cursor == '':
            break
            

def get_eth_block(api_link):
    tx = requests.get(api_link)
    content = json.loads(tx.content)
    return int(content['result'][0]['blockNumber'])

def process_scan(asset, asset_symbol, asset_decimal, address, start_date, end_date, output):
    domain = 'api.etherscan.io'
    api_key = 'XQBQNA7X7Z8BU41N9DZ5WUGT5E7YSFF15M'
    tx_url = 'https://etherscan.io/tx/'

    if asset_symbol == 'BSC':
        domain = 'api.bscscan.com'
        api_key = 'IG49A7A5VG6SEA2ZBSDEH2NXRHTNDN93B2'
        tx_url = 'https://bscscan.com/tx/'
    elif asset_symbol == 'MATIC':
        domain = 'api.polygonscan.com'
        api_key = 'FD9PJJDSJ6JX3K2ZKWJIWRJMEWXEBG31QU'
        tx_url = 'https://polygonscan.com/tx/'

    api_link = f'https://{domain}/api?module=account&address={address}&apikey={api_key}'
    
    
    address = address.lower()
    
    work_sheet = set_worksheet(output, asset)
    work_sheet.append(['Date', 'Block Index', 'Transaction ID', 'Sent', 'Received', 'Asset', 'Coming From', 'Going To', 'Blockchain URL']) 
    
    start_block = get_eth_block(f'{api_link}&action=txlist&startblock=0&endblock=99999999&sort=asc&offset=1&page=1')
    end_block = get_eth_block(f'{api_link}&action=txlist&startblock=0&endblock=99999999&sort=desc&offset=1&page=1')
    
    block_increment = 100000
    while start_block <= end_block:
        print(f'Working on blocks {start_block}-{start_block + block_increment}')
        api_txlist = f'{api_link}&action=txlist&startblock={start_block}&endblock={start_block + block_increment}&sort=asc'
        api_tokentx = f'{api_link}&action=tokentx&startblock={start_block}&endblock={start_block + block_increment}&sort=asc'

        tx_list = requests.get(api_txlist)
        content = json.loads(tx_list.content)
        rows = content['result']
        for row in rows:
            timestamp = int(row['timeStamp'])
            if timestamp < start_date:
                continue
            if timestamp > end_date:
                break
                
            value = Decimal(row['value'])
            if value > 0:
                value /= Decimal(asset_decimal)
                sent = 0
                received = 0
                if address == row['from']:
                    sent = value
                elif address == row['to']:
                    received = value
                    
                work_sheet.append([str(datetime.fromtimestamp(timestamp)), row['blockNumber'], row['hash'], sent, received, asset_symbol, row['from'], row['to'], ''])
                set_verified_tx(work_sheet, work_sheet.max_row, 8, f'{tx_url}{row["hash"]}')

        token_tx = requests.get(api_tokentx)
        content = json.loads(token_tx.content)
        rows = content['result']
        for row in rows:
            token_symbol = row['tokenSymbol']
        
            timestamp = int(row['timeStamp'])
            if timestamp < start_date:
                continue
            if timestamp > end_date:
                break
                
            value = Decimal(row['value']) / Decimal(10 ** int(row['tokenDecimal']))
            sent = 0
            received = 0
            if address == row['from']:
                sent = value
            elif address == row['to']:
                received = value
                    
            work_sheet.append([str(datetime.fromtimestamp(timestamp)), row['blockNumber'], row['hash'], sent, received, token_symbol, row['from'], row['to'], ''])
            set_verified_tx(work_sheet, work_sheet.max_row, 8, f'{tx_url}{row["hash"]}')
        
        start_block += block_increment + 1
        
                
def process_blockchair_asset(asset, asset_symbol, asset_decimal, api_keyword, address, start_date, end_date, output):
    start_date = datetime.fromtimestamp(start_date).strftime('%Y-%m-%d')
    end_date = datetime.fromtimestamp(end_date).strftime('%Y-%m-%d')
    api_key = 'change it'
    txns = requests.get(f'https://api.blockchair.com/{api_keyword}/dashboards/address/{address}?transaction_details=true&q=time({start_date}..{end_date})&key={api_key}')
   
    content = json.loads(txns.content)
   
    if content['data'] is None or content['data'][address]['address']['type'] is None:
        return
       
    work_sheet = set_worksheet(output, asset)
    work_sheet.append(['Date', 'Block Index', 'Transaction ID', 'Sent', 'Received', 'Asset', 'Blockchain URL'])
   
    rows = content['data'][address]['transactions']
   
    for row in rows:
        value = int(row['balance_change'])
        sent = value*-1/asset_decimal if value < 0 else 0
        received = value/asset_decimal if value > 0 else 0
        
       
        work_sheet.append([row['time'], row['block_id'], row['hash'], sent, received, asset_symbol, ''])
        set_verified_tx(work_sheet, work_sheet.max_row, 6, f'{ASSET_TX_URLS[asset_symbol]}{row["hash"]}')    


        
def process_blockdaemon_asset_without_date_filter(asset, asset_symbol, asset_decimal, api_keyword, address, start_date, end_date, output):
    api_link = f'https://svc.blockdaemon.com/universal/v1/{api_keyword}/mainnet/account/{address}/txs?page_size=100&order=asc'
    txs = requests.get(api_link, headers = {"Authorization": f"Bearer {BLOCKDAEMON_BEARER}"})
    content = json.loads(txs.content)

    work_sheet = set_worksheet(output, asset)
    work_sheet.append(['Date', 'Block Index', 'Transaction ID', 'Sent', 'Received', 'Asset', 'Coming From', 'Going To', 'Blockchain URL'])
    
    while content:
        for row in content['data']:
            timestamp = int(row['date'])
            if timestamp < start_date:
                continue
            if timestamp > end_date:
                break
                
            date = str(datetime.fromtimestamp(timestamp))     
            sent = 0
            received = 0
            for event in row['events']:
                if 'destination' in event and  event['destination'] == address:
                    received += event['amount']
                elif 'source' in event and event['source'] == address:
                    sent += event['amount']
                    
            if sent:
                sent /= asset_decimal
            if received:
                received /= asset_decimal
        
            work_sheet.append([date, row['block_number'], row['id'], sent, received, asset_symbol, event['source'], event['destination'], ''])
            set_verified_tx(work_sheet, work_sheet.max_row, 8, f'{ASSET_TX_URLS[asset_symbol]}{row["id"]}')        
            
        try:
            next_page_token = content['meta']['paging']['next_page_token']
        except:
            next_page_token = None
            
            
        if next_page_token:
            next_link = f'{api_link}&page_token={next_page_token}'
            txs = requests.get(next_link, headers = {"Authorization": f"Bearer {BLOCKDAEMON_BEARER}"})
            content = json.loads(txs.content)
        else:
            content = None        


def process_blockdaemon_asset(asset, asset_symbol, asset_decimal, api_keyword, address, start_date, end_date, output):
    api_link = f'https://svc.blockdaemon.com/universal/v1/{api_keyword}/mainnet/account/{address}/txs?page_size=100&from={start_date}&to={end_date}'
    txs = requests.get(api_link, headers = {"Authorization": f"Bearer {BLOCKDAEMON_BEARER}"})
    content = json.loads(txs.content)

    work_sheet = set_worksheet(output, asset)
    work_sheet.append(['Date', 'Block Index', 'Transaction ID', 'Sent', 'Received', 'Asset', 'Coming From', 'Going To', 'Blockchain URL'])
    
    while content:
        for row in content['data']:
            timestamp = int(row['date'])
            date = str(datetime.fromtimestamp(timestamp))     
            sent = 0
            received = 0
            for event in row['events']:
                if 'destination' in event and  event['destination'] == address:
                    received += event['amount']
                elif 'source' in event and event['source'] == address:
                    sent += event['amount']
                    
            if sent:
                sent /= asset_decimal
            if received:
                received /= asset_decimal
        
            work_sheet.append([date, row['block_number'], row['id'], sent, received, asset_symbol, event['source'], event['destination'], ''])
            set_verified_tx(work_sheet, work_sheet.max_row, 8, f'{ASSET_TX_URLS[asset_symbol]}{row["id"]}')        
            
        try:
            next_page_token = content['meta']['paging']['next_page_token']
        except:
            next_page_token = None
            
            
        if next_page_token:
            next_link = f'{api_link}&page_token={next_page_token}'
            txs = requests.get(next_link, headers = {"Authorization": f"Bearer {BLOCKDAEMON_BEARER}"})
            content = json.loads(txs.content)
        else:
            content = None
             
def process_address(asset, address, start_date, end_date, output):
    if asset is None:
        return
        
    if asset.lower() == 'bitcoin':
        process_blockchair_asset(asset, 'BTC', 1e8, 'bitcoin', address, start_date, end_date, output)
    elif asset.lower() == 'bitcoin cash':
        process_blockchair_asset(asset, 'BCH', 1e8, 'bitcoin-cash', address, start_date, end_date, output)
    elif asset.lower() == 'dogechain':
        process_blockchair_asset(asset, 'DOGE', 1e8, 'dogecoin', address, start_date, end_date, output)
    elif asset.lower() == 'dash':
        process_blockchair_asset(asset, 'DASH', 1e8, 'dash', address, start_date, end_date, output)
    elif asset.lower() == 'litecoin':
        process_blockchair_asset(asset, 'LTC', 1e8, 'litecoin', address, start_date, end_date, output)
    elif asset.lower() == 'zcash':
        process_blockchair_asset(asset, 'ZEC', 1e8, 'zcash', address, start_date, end_date, output)
    elif asset.lower() == 'ethereum':
       process_scan(asset, 'ETH', 1e18, address, start_date, end_date, output)
    elif asset.lower() == 'binance smart chain':
        process_scan(asset, 'BSC', 1e18, address, start_date, end_date, output)
    elif asset.lower() == 'polygon':
        process_scan(asset, 'MATIC', 1e18, address, start_date, end_date, output)
    elif asset.lower() == 'polkadot':
       process_blockdaemon_asset(asset, 'DOT', 1e10, 'polkadot', address, start_date, end_date, output)
    elif asset.lower() == 'solana':
        process_blockdaemon_asset(asset, 'SOL', 1e9, 'solana', address, start_date, end_date, output)
    elif asset.lower() == 'algorand':
        process_blockdaemon_asset(asset, 'ALGO', 1e6, 'algorand', address, start_date, end_date, output)
    elif asset.lower() == 'ripple':
        process_blockdaemon_asset_without_date_filter(asset, 'XRP', 1e6, 'xrp', address, start_date, end_date, output)

    

def process_input(asset, address, start_date, end_date, output):
    print(f'Naked Adress running... {datetime.now()}')
    print(f'Working with address {address}') 
    process_address(asset, address, start_date, end_date, output)            
    print(f'Processing Done. {datetime.now()}')
    

def get_history(asset, before, after):
    history = requests.get(f'https://api.cryptowat.ch/markets/binance-us/{asset}/ohlc?before={before}&after={after}&periods=86400', headers=HEADERS)
    content = json.loads(history.content)
    return content['result']['86400']

def generate_historical_price_data(start_date, end_date, asset):    
    rows = get_history(asset, end_date + 28800, start_date - 28800)
    
    filename = f'{OUTPUT_PATH}{asset}_price_history_{start_date}-{end_date}.xlsx'
    
    work_book = Workbook()
    work_sheet = work_book.active
    work_sheet.append(['Date', 'Value(USD)'])

    for row in rows:
        date = datetime.fromtimestamp(int(row[0])).strftime('%m/%d/%Y')
        work_sheet.append([date, row[4]])        
    
    work_book.save(filename)


def run():
    print('nakedAddress (terminalWallet) will ask you to add the type of network, wallet address and your start/end dates. Type your response and press enter.')
    print('For price historical data, type "price" where it asks for wallet address')
    print('Write the complete name for the type of network, for example, Ethereum, not ETH')
    print('Available Networks: Algorand, Avalanche, Binance Smart Chain, Bitcoin, Bitcoin Cash, Cardano, Dash, Dogechain, Ethereum, Litecoin, Polygon, Ripple, Zcash.')
    asset = input('Network:').strip()
    address = input('Address:').strip()
    start_date = int(datetime.strptime(input('Start Date (m/d/yyyy): '), "%m/%d/%Y").timestamp())
    end_date = int(datetime.strptime(input('End Date (m/d/yyyy): '), "%m/%d/%Y").timestamp())
    
    if address == 'price':
        print(f'Generating price history for {asset}...')
        generate_historical_price_data(start_date, end_date, asset_usd_mapping[asset])
        return
    
    output = Workbook()
    process_input(asset, address, start_date, end_date, output)
    output.save(f'{OUTPUT_PATH}{address}.xlsx')
